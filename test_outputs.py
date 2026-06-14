import os
from pathlib import Path

import openpyxl


OUTPUT_FILE = Path(os.getenv("OUTPUT_FILE", "careflow_accessibility_triage.xlsx"))


def load_workbook():
    assert OUTPUT_FILE.exists(), f"Missing output file: {OUTPUT_FILE}"
    return openpyxl.load_workbook(OUTPUT_FILE, data_only=True)


def workbook_text(wb):
    values = []
    for sheet in wb.worksheets:
        values.append(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
    return " ".join(values)


def sheet_text(wb, sheet_name):
    assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"
    values = []
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                values.append(str(cell))
    return " ".join(values)


def test_required_file_and_sheets_exist():
    wb = load_workbook()
    expected_sheets = [
        "Intake Summary",
        "Visual Blockers",
        "Tracking Noise",
        "Final Review",
    ]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames


def test_intake_summary_values():
    wb = load_workbook()
    text = sheet_text(wb, "Intake Summary")
    assert "Chad Ramsey" in text
    assert "CareFlow Intake" in text
    assert "design review blocker check" in text
    assert "careflow_accessibility_triage.xlsx" in text
    assert "4" in text
    assert "3" in text
    assert "1" in text
    assert "ready for design review" in text


def test_visual_blockers_include_current_issue_ids():
    wb = load_workbook()
    text = sheet_text(wb, "Visual Blockers")
    for issue_id in ["INT-014", "INT-018", "INT-021", "INT-024"]:
        assert issue_id in text
    assert "REF-021" not in text
    assert "INT-018B" not in text


def test_visual_blocker_details_are_correct():
    wb = load_workbook()
    text = sheet_text(wb, "Visual Blockers")
    assert "low contrast primary action" in text
    assert "2.9" in text
    assert "4.5" in text
    assert "focus order skips error summary" in text
    assert "after field group" in text
    assert "before field group" in text
    assert "mobile tap target below 44px" in text
    assert "36" in text
    assert "44" in text
    assert "select field missing accessible name" in text
    assert "Choose option" in text
    assert "Insurance plan type" in text


def test_tracking_noise_is_separated():
    wb = load_workbook()
    text = sheet_text(wb, "Tracking Noise")
    assert "REF-021" in text
    assert "CareFlow Referral" in text
    assert "INT-018B" in text
    assert "stale duplicate" in text
    assert "old_audit_v3.pdf" in text
    assert "stale" in text


def test_final_review_decision():
    wb = load_workbook()
    text = sheet_text(wb, "Final Review")
    assert "CareFlow Intake" in text
    assert "4" in text
    assert "INT-014" in text
    assert "INT-018" in text
    assert "INT-021" in text
    assert "INT-024" in text
    assert "REF-021" in text
    assert "INT-018B" in text
    assert "No message" in text or "no message" in text
